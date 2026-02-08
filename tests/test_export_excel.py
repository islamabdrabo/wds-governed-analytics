from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def load_export_module():
    module_path = Path(__file__).resolve().parents[1] / "import" / "06_export_excel.py"
    spec = spec_from_file_location("export_module_for_tests", module_path)
    module = module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def test_export_official_excel_includes_original_csv_drilldown(tmp_path):
    module = load_export_module()

    source_csv = tmp_path / "source.csv"
    source_df = pd.DataFrame(
        [
            {
                "civil id": "1",
                "file number": "10",
                "full name": "Alice",
                "nationality": "A",
                "job title": "X",
                "region": "R1",
                "workplace": "W1",
                "education level": "E1",
                "Specialization in study": "S1",
                "supervisory title": "SV1",
                "hire date": "2020-01-01",
                "pay grade": "G1",
                "final specialty": "F1",
            },
            {
                "civil id": "2",
                "file number": "20",
                "full name": "Bob",
                "nationality": "B",
                "job title": "Y",
                "region": "R2",
                "workplace": "W2",
                "education level": "E2",
                "Specialization in study": "S2",
                "supervisory title": "SV2",
                "hire date": "2021-01-01",
                "pay grade": "G2",
                "final specialty": "F2",
            },
        ]
    )
    source_df.to_csv(source_csv, index=False)

    # Redirect drilldown source for this test.
    module.CSV_PATH = source_csv

    base_df = pd.DataFrame(
        [
            {"person_id": "P1", "region_name": "R1", "workplace_name": "W1", "specialty_name": "F1"},
            {"person_id": "P2", "region_name": "R2", "workplace_name": "W2", "specialty_name": "F2"},
        ]
    )

    output_path = tmp_path / "out.xlsx"
    module.export_official_excel(base_df, output_path)

    wb = load_workbook(output_path, read_only=False)
    assert wb.sheetnames == [
        "Region x Specialty",
        "Region+Workplace x Specialty",
        "Drilldown_Filtered",
    ]

    ws_region = wb["Region x Specialty"]
    ws_region_wp = wb["Region+Workplace x Specialty"]
    ws_drilldown = wb["Drilldown_Filtered"]

    assert ws_region.freeze_panes == "B2"
    assert ws_region_wp.freeze_panes == "C2"

    drilldown_headers = [ws_drilldown.cell(1, c).value for c in range(1, ws_drilldown.max_column + 1)]
    assert drilldown_headers == list(source_df.columns)
    assert ws_drilldown.max_row == len(source_df) + 1


def test_export_official_excel_validates_required_columns(tmp_path):
    module = load_export_module()
    output_path = tmp_path / "out.xlsx"

    invalid_df = pd.DataFrame([{"person_id": "P1", "region_name": "R1"}])
    try:
        module.export_official_excel(invalid_df, output_path)
    except ValueError as exc:
        assert "Missing required columns" in str(exc)
    else:
        raise AssertionError("Expected ValueError for missing required columns")
