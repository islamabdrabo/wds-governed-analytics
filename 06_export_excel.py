from __future__ import annotations

import sqlite3
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Iterable, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config.paths import CSV_PATH, DB_PATH, EXPORTS_DIR

REQUIRED_COLUMNS = {
    "person_id",
    "region_name",
    "workplace_name",
    "specialty_name",
}

TOTAL_LABEL = "الإجمالي"

OutputTarget = Union[str, Path, BytesIO, BinaryIO]


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF404040")
BODY_FONT = Font(name="Calibri", size=11, bold=False)
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
ROW_FILLS = [
    PatternFill(fill_type="solid", fgColor="FFE9EDF3"),  # style 3
    PatternFill(fill_type="solid", fgColor="FFDDE5ED"),  # style 5
    PatternFill(fill_type="solid", fgColor="FFF2EEE8"),  # style 6
    PatternFill(fill_type="solid", fgColor="FFE6EBE7"),  # style 7
    PatternFill(fill_type="solid", fgColor="FFEEF1F4"),  # style 8
]
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_ALIGN = Alignment(horizontal="center", vertical="top")


def _validate_columns(df: pd.DataFrame) -> None:
    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns for export: {sorted(missing)}")


def _build_pivots(df_base: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    grp_region = (
        df_base.groupby(["region_name", "specialty_name"], dropna=False)
        .size()
        .reset_index(name="count")
    )
    pivot_region = grp_region.pivot_table(
        index="region_name",
        columns="specialty_name",
        values="count",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name=TOTAL_LABEL,
    )
    pivot_region = pivot_region.reset_index()

    grp_region_wp = (
        df_base.groupby(
            ["region_name", "workplace_name", "specialty_name"],
            dropna=False,
        )
        .size()
        .reset_index(name="count")
    )
    pivot_region_wp = grp_region_wp.pivot_table(
        index=["region_name", "workplace_name"],
        columns="specialty_name",
        values="count",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name=TOTAL_LABEL,
    )
    pivot_region_wp = pivot_region_wp.reset_index()

    return pivot_region, pivot_region_wp


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    clean = df.copy()
    clean["person_id"] = clean["person_id"].fillna("").astype(str).str.strip()
    clean["region_name"] = clean["region_name"].fillna("").astype(str).str.strip()
    clean["workplace_name"] = clean["workplace_name"].fillna("").astype(str).str.strip()
    clean["specialty_name"] = clean["specialty_name"].fillna("").astype(str).str.strip()
    return clean


def _safe_width(value: object) -> int:
    text = str(value) if value is not None else ""
    return max(7, min(45, len(text) + 2))


def _load_original_drilldown_data() -> pd.DataFrame:
    if not CSV_PATH.exists():
        raise FileNotFoundError(f"Source file not found: {CSV_PATH}")

    try:
        return pd.read_csv(CSV_PATH, low_memory=False)
    except UnicodeDecodeError:
        return pd.read_csv(CSV_PATH, encoding="utf-8-sig", low_memory=False)


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    headers = [str(c) for c in df.columns]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _style_header_row(ws, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _style_total_row(ws, row_idx: int, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _style_region_sheet(ws, data_row_count: int, col_count: int) -> None:
    _style_header_row(ws, col_count)

    total_row = data_row_count + 1
    for row_idx in range(2, total_row):
        row_fill = ROW_FILLS[(row_idx - 2) % len(ROW_FILLS)]
        for col_idx in range(1, col_count):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

    _style_total_row(ws, total_row, col_count)

    # Total column is always gray/bold for all data rows.
    for row_idx in range(2, total_row):
        cell = ws.cell(row=row_idx, column=col_count)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    ws.freeze_panes = "B2"

    ws.column_dimensions["A"].width = 31
    for col_idx in range(2, col_count + 1):
        col_values = [ws.cell(row=1, column=col_idx).value]
        col_values.extend(
            ws.cell(row=row_idx, column=col_idx).value
            for row_idx in range(2, min(ws.max_row, 30) + 1)
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            _safe_width(v) for v in col_values
        )


def _style_region_workplace_sheet(ws, data_row_count: int, col_count: int) -> None:
    _style_header_row(ws, col_count)

    total_row = data_row_count + 1
    current_region = None
    style_index = -1

    for row_idx in range(2, total_row):
        region = ws.cell(row=row_idx, column=1).value
        if region != current_region:
            current_region = region
            style_index = (style_index + 1) % len(ROW_FILLS)
        row_fill = ROW_FILLS[style_index]

        for col_idx in range(1, col_count):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

    _style_total_row(ws, total_row, col_count)

    for row_idx in range(2, total_row):
        cell = ws.cell(row=row_idx, column=col_count)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    ws.freeze_panes = "C2"

    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 45
    for col_idx in range(3, col_count + 1):
        col_values = [ws.cell(row=1, column=col_idx).value]
        col_values.extend(
            ws.cell(row=row_idx, column=col_idx).value
            for row_idx in range(2, min(ws.max_row, 30) + 1)
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            _safe_width(v) for v in col_values
        )


def _style_drilldown_sheet(ws, df: pd.DataFrame) -> None:
    col_count = len(df.columns)

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGN

    for col_idx in range(1, col_count + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        header = ws.cell(row=1, column=col_idx).value
        max_data = ""
        if ws.max_row > 1:
            max_data = max(
                (str(ws.cell(row=row_idx, column=col_idx).value or "") for row_idx in range(2, ws.max_row + 1)),
                key=len,
                default="",
            )
        ws.column_dimensions[col_letter].width = max(
            10,
            min(50, max(len(str(header or "")), len(max_data)) + 2),
        )


def export_official_excel(df_base: pd.DataFrame, output: OutputTarget) -> OutputTarget:
    """
    Export a filtered canonical dataframe to the official workbook format.
    """
    _validate_columns(df_base)
    if df_base.empty:
        raise ValueError("No data available for export.")

    records_df = _sanitize_df(df_base)
    pivot_region, pivot_region_wp = _build_pivots(records_df)
    drilldown_df = _load_original_drilldown_data()

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_region = wb.create_sheet("Region x Specialty")
    _write_dataframe(ws_region, pivot_region)
    _style_region_sheet(ws_region, data_row_count=len(pivot_region), col_count=len(pivot_region.columns))

    ws_region_wp = wb.create_sheet("Region+Workplace x Specialty")
    _write_dataframe(ws_region_wp, pivot_region_wp)
    _style_region_workplace_sheet(
        ws_region_wp,
        data_row_count=len(pivot_region_wp),
        col_count=len(pivot_region_wp.columns),
    )

    ws_drilldown = wb.create_sheet("Drilldown_Filtered")
    _write_dataframe(ws_drilldown, drilldown_df)
    _style_drilldown_sheet(ws_drilldown, drilldown_df)

    wb.save(output)
    return output


def _load_filtered_rows(
    selected_regions: Iterable[str] | None = None,
    selected_workplaces: Iterable[str] | None = None,
    selected_specialties: Iterable[str] | None = None,
) -> pd.DataFrame:
    selected_regions = list(selected_regions or [])
    selected_workplaces = list(selected_workplaces or [])
    selected_specialties = list(selected_specialties or [])

    query = """
        SELECT
            person_id,
            region_name,
            workplace_name,
            specialty_name
        FROM v_workforce_base_canonical
        WHERE 1 = 1
    """
    params: list[str] = []

    if selected_regions:
        placeholders = ",".join("?" for _ in selected_regions)
        query += f" AND region_name IN ({placeholders})"
        params.extend(selected_regions)

    if selected_workplaces:
        placeholders = ",".join("?" for _ in selected_workplaces)
        query += f" AND workplace_name IN ({placeholders})"
        params.extend(selected_workplaces)

    if selected_specialties:
        placeholders = ",".join("?" for _ in selected_specialties)
        query += f" AND specialty_name IN ({placeholders})"
        params.extend(selected_specialties)

    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql(query, conn, params=params)
    finally:
        conn.close()


def export_workforce_excel(
    selected_regions: Iterable[str] | None = None,
    selected_workplaces: Iterable[str] | None = None,
    selected_specialties: Iterable[str] | None = None,
    output_filename: str = "Workforce_Analytics.xlsx",
) -> Path:
    """
    Backward-compatible helper that reads from DB then exports to disk.
    """
    df_base = _load_filtered_rows(
        selected_regions=selected_regions,
        selected_workplaces=selected_workplaces,
        selected_specialties=selected_specialties,
    )
    if df_base.empty:
        raise ValueError("No data available for the selected filters.")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = EXPORTS_DIR / output_filename
    export_official_excel(df_base, output_path)
    return output_path
